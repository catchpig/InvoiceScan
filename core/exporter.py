from enum import Enum
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from models.invoice import Invoice, InvoiceItem

_MIN_COL_WIDTH = 10
_MAX_COL_WIDTH = 50
_COL_PADDING = 2


class ExportMode(Enum):
    SUMMARY = "summary"
    DETAIL = "detail"


_SUMMARY_HEADERS = [
    "来源文件", "发票类型", "发票代码", "发票号码", "开票日期",
    "购买方名称", "购买方税号", "销售方名称", "销售方税号",
    "不含税金额", "税率", "税额", "价税合计", "开票人", "状态",
]

_DETAIL_HEADERS = [
    "来源文件", "发票代码", "发票号码", "开票日期",
    "购买方名称", "销售方名称",
    "货物/服务名称", "数量", "单价", "金额",
    "税率", "税额", "价税合计",
]


class Exporter:
    def export(self, invoices: list[Invoice], output_path: str,
               mode: ExportMode, dedup: bool = True) -> int:
        """Export invoices to Excel.

        Args:
            dedup: When True, keep only the last occurrence of invoices
                   sharing the same (invoice_code, invoice_number).

        Returns:
            Number of duplicate invoices removed.
        """
        removed = 0
        if dedup:
            deduped, removed = self._deduplicate(invoices)
            invoices = deduped

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "汇总" if mode == ExportMode.SUMMARY else "明细"

        if mode == ExportMode.SUMMARY:
            self._write_summary(ws, invoices)
        else:
            self._write_detail(ws, invoices)

        wb.save(output_path)
        return removed

    @staticmethod
    def _deduplicate(invoices: list[Invoice]) -> tuple[list[Invoice], int]:
        """Keep last occurrence when invoices are considered duplicates.

        Dedup keys (tried in order, first non-empty key wins):
        1. invoice_code + invoice_number (primary)
        2. invoice_number only (if code is empty but number exists)
        3. source_file (fallback — same filename likely same invoice)

        Returns:
            (deduplicated list, number of removed duplicates)
        """
        def _dedup_key(inv: Invoice) -> tuple | None:
            """Return a dedup key for the invoice, or None if no key possible."""
            if inv.invoice_code and inv.invoice_number:
                return ("code_num", inv.invoice_code, inv.invoice_number)
            if inv.invoice_number:
                return ("num", inv.invoice_number)
            if inv.source_file:
                return ("file", inv.source_file)
            return None

        # Find the last index for each key
        last_index: dict[tuple, int] = {}
        for i, inv in enumerate(invoices):
            key = _dedup_key(inv)
            if key is not None:
                last_index[key] = i  # later index overwrites earlier

        result = []
        for i, inv in enumerate(invoices):
            key = _dedup_key(inv)
            if key is None:
                result.append(inv)  # no key → always keep
            elif last_index[key] == i:
                result.append(inv)  # keep only the last occurrence
        return result, len(invoices) - len(result)

    def _write_summary(self, ws, invoices: list[Invoice]) -> None:
        self._write_header(ws, _SUMMARY_HEADERS)
        rows = [
            [
                inv.source_file, inv.invoice_type,
                inv.invoice_code, inv.invoice_number, inv.invoice_date,
                inv.buyer_name, inv.buyer_tax_id,
                inv.seller_name, inv.seller_tax_id,
                str(inv.subtotal), inv.tax_rate, str(inv.tax_amount),
                str(inv.total_amount), inv.issuer, inv.status,
            ]
            for inv in invoices
        ]
        for row in rows:
            ws.append(row)
        self._apply_column_widths(ws, self._calc_column_widths(_SUMMARY_HEADERS, rows))
        ws.freeze_panes = "A2"

    def _write_detail(self, ws, invoices: list[Invoice]) -> None:
        self._write_header(ws, _DETAIL_HEADERS)
        rows = []
        for inv in invoices:
            items = inv.items if inv.items else [None]
            for item in items:
                rows.append([
                    inv.source_file, inv.invoice_code, inv.invoice_number,
                    inv.invoice_date, inv.buyer_name, inv.seller_name,
                    item.name if item else "",
                    item.quantity if item else "",
                    str(item.unit_price) if item else "",
                    str(item.amount) if item else "",
                    inv.tax_rate, str(inv.tax_amount), str(inv.total_amount),
                ])
        for row in rows:
            ws.append(row)
        self._apply_column_widths(ws, self._calc_column_widths(_DETAIL_HEADERS, rows))
        ws.freeze_panes = "A2"

    def _write_header(self, ws, headers: list[str]) -> None:
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

    @staticmethod
    def _display_width(text) -> float:
        return sum(2.0 if ord(c) > 127 else 1.0 for c in str(text or ""))

    @classmethod
    def _calc_column_widths(cls, headers: list[str], rows: list[list]) -> list[float]:
        widths = [cls._display_width(h) for h in headers]
        for row in rows:
            for i, val in enumerate(row):
                w = cls._display_width(val)
                if w > widths[i]:
                    widths[i] = w
        return [max(_MIN_COL_WIDTH, min(_MAX_COL_WIDTH, w + _COL_PADDING)) for w in widths]

    @staticmethod
    def _apply_column_widths(ws, widths: list[float]) -> None:
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width
