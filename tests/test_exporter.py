import os
import tempfile
import openpyxl
from decimal import Decimal
from openpyxl.utils import get_column_letter
from core.exporter import Exporter, ExportMode, _MIN_COL_WIDTH, _MAX_COL_WIDTH
from models.invoice import Invoice, InvoiceItem, InvoiceStatus


def _make_invoice(code="044001900111", number="12345678") -> Invoice:
    return Invoice(
        source_file="test.pdf",
        invoice_code=code,
        invoice_number=number,
        invoice_date="2024-03-15",
        buyer_name="深圳XX科技有限公司",
        buyer_tax_id="914403001234567890",
        seller_name="广州YY贸易有限公司",
        seller_tax_id="914401011234567891",
        items=[InvoiceItem(name="软件开发服务", quantity="1",
                           unit_price=Decimal("1000.00"), amount=Decimal("1000.00"))],
        subtotal=Decimal("1000.00"),
        tax_rate="13%",
        tax_amount=Decimal("130.00"),
        total_amount=Decimal("1130.00"),
        issuer="张三",
        invoice_type="增值税专用发票",
        status=InvoiceStatus.SUCCESS,
    )


def test_export_summary_creates_excel():
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        path = f.name
    try:
        Exporter().export([_make_invoice()], path, ExportMode.SUMMARY)
        wb = openpyxl.load_workbook(path)
        assert wb.active.max_row == 2  # 表头 + 1 行数据
    finally:
        os.unlink(path)


def test_export_summary_has_required_headers():
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        path = f.name
    try:
        Exporter().export([_make_invoice()], path, ExportMode.SUMMARY)
        ws = openpyxl.load_workbook(path).active
        headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
        for required in ["发票代码", "发票号码", "价税合计", "购买方名称", "销售方名称"]:
            assert required in headers
    finally:
        os.unlink(path)


def test_export_summary_data_row_values():
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        path = f.name
    try:
        Exporter().export([_make_invoice("044001900111", "12345678")], path, ExportMode.SUMMARY)
        ws = openpyxl.load_workbook(path).active
        row2 = [ws.cell(2, col).value for col in range(1, ws.max_column + 1)]
        assert "044001900111" in row2
        assert "12345678" in row2
    finally:
        os.unlink(path)


def test_export_detail_expands_items():
    inv = _make_invoice()
    inv.items = [
        InvoiceItem(name="服务A", quantity="1", unit_price=Decimal("500"), amount=Decimal("500")),
        InvoiceItem(name="服务B", quantity="2", unit_price=Decimal("250"), amount=Decimal("500")),
    ]
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        path = f.name
    try:
        Exporter().export([inv], path, ExportMode.DETAIL)
        ws = openpyxl.load_workbook(path).active
        assert ws.max_row == 3  # 表头 + 2 条明细行
    finally:
        os.unlink(path)


def test_export_multiple_invoices_summary():
    invoices = [_make_invoice("111", "001"), _make_invoice("222", "002")]
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        path = f.name
    try:
        Exporter().export(invoices, path, ExportMode.SUMMARY)
        assert openpyxl.load_workbook(path).active.max_row == 3
    finally:
        os.unlink(path)


def test_export_invoice_without_items_detail():
    inv = _make_invoice()
    inv.items = []
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        path = f.name
    try:
        Exporter().export([inv], path, ExportMode.DETAIL)
        assert openpyxl.load_workbook(path).active.max_row == 2
    finally:
        os.unlink(path)


# --- 列宽相关测试 ---

def test_display_width_accuracy():
    assert Exporter._display_width("abc") == 3.0
    assert Exporter._display_width("购买方") == 6.0
    assert Exporter._display_width("INV001 发票") == 11.0
    assert Exporter._display_width("") == 0.0
    assert Exporter._display_width(None) == 0.0


def test_column_widths_expand_for_long_text():
    inv = _make_invoice()
    inv.buyer_name = "深圳市超长示例公司名称科技发展有限责任公司"  # 15 个汉字 = 显示宽度 30
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        path = f.name
    try:
        Exporter().export([inv], path, ExportMode.SUMMARY)
        ws = openpyxl.load_workbook(path).active
        buyer_col = get_column_letter(6)  # 购买方名称是第 6 列
        assert ws.column_dimensions[buyer_col].width > Exporter._display_width("购买方名称")
    finally:
        os.unlink(path)


def test_column_widths_capped_at_max():
    inv = _make_invoice()
    inv.source_file = "C:\\" + "a" * 200 + ".pdf"  # 超长路径
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        path = f.name
    try:
        Exporter().export([inv], path, ExportMode.SUMMARY)
        ws = openpyxl.load_workbook(path).active
        assert ws.column_dimensions[get_column_letter(1)].width <= _MAX_COL_WIDTH
    finally:
        os.unlink(path)


def test_all_columns_have_min_width():
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        path = f.name
    try:
        Exporter().export([_make_invoice()], path, ExportMode.SUMMARY)
        ws = openpyxl.load_workbook(path).active
        for col_idx in range(1, ws.max_column + 1):
            assert ws.column_dimensions[get_column_letter(col_idx)].width >= _MIN_COL_WIDTH
    finally:
        os.unlink(path)


# --- 去重相关测试 ---

def test_deduplicate_keeps_last_occurrence():
    """相同发票代码+号码出现多次时，保留最后一条。"""
    inv1 = _make_invoice("A001", "10001")
    inv1.buyer_name = "第一次"
    inv2 = _make_invoice("A001", "10001")
    inv2.buyer_name = "第二次"
    inv3 = _make_invoice("B002", "20002")
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        path = f.name
    try:
        removed = Exporter().export([inv1, inv2, inv3], path, ExportMode.SUMMARY)
        ws = openpyxl.load_workbook(path).active
        # 2 unique invoices (A001 deduped to 1 + B002 = 2) → header + 2 rows
        assert ws.max_row == 3
        assert removed == 1
        # The remaining A001 row should be "第二次" (the last one)
        buyer_col = 6
        a001_row = [ws.cell(r, buyer_col).value for r in range(2, ws.max_row + 1)
                    if ws.cell(r, 3).value == "A001"]
        assert a001_row == ["第二次"]
    finally:
        os.unlink(path)


def test_deduplicate_preserves_order():
    """去重后保持原始顺序。"""
    inv1 = _make_invoice("A001", "10001")
    inv2 = _make_invoice("B002", "20002")
    inv3 = _make_invoice("A001", "10001")  # duplicate of inv1
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        path = f.name
    try:
        Exporter().export([inv1, inv2, inv3], path, ExportMode.SUMMARY)
        ws = openpyxl.load_workbook(path).active
        codes = [ws.cell(r, 3).value for r in range(2, ws.max_row + 1)]
        assert codes == ["B002", "A001"]  # A001 kept at its last position
    finally:
        os.unlink(path)


def test_deduplicate_keeps_entries_without_code_or_number():
    """发票代码和号码都为空时，按 source_file fallback 去重。"""
    inv1 = _make_invoice("A001", "10001")
    inv2 = Invoice(source_file="unknown.pdf", invoice_code="", invoice_number="",
                   buyer_name="无码发票1")
    inv3 = Invoice(source_file="unknown.pdf", invoice_code="", invoice_number="",
                   buyer_name="无码发票2")
    inv4 = _make_invoice("A001", "10001")  # duplicate of inv1
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        path = f.name
    try:
        removed = Exporter().export([inv1, inv2, inv3, inv4], path, ExportMode.SUMMARY)
        ws = openpyxl.load_workbook(path).active
        # inv1 deduped (code+num), inv2 deduped by source_file fallback, inv3+inv4 kept → 2 data rows
        assert ws.max_row == 3
        assert removed == 2
    finally:
        os.unlink(path)


def test_deduplicate_by_number_only():
    """发票代码为空但号码相同时，按号码去重。"""
    inv1 = Invoice(source_file="a.pdf", invoice_code="", invoice_number="10001",
                   buyer_name="第一次")
    inv2 = Invoice(source_file="b.pdf", invoice_code="", invoice_number="10001",
                   buyer_name="第二次")
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        path = f.name
    try:
        removed = Exporter().export([inv1, inv2], path, ExportMode.SUMMARY)
        ws = openpyxl.load_workbook(path).active
        assert ws.max_row == 2  # header + 1 row
        assert removed == 1
        # Kept the last one
        buyer_col = 6
        assert ws.cell(2, buyer_col).value == "第二次"
    finally:
        os.unlink(path)


def test_deduplicate_disabled():
    """dedup=False 时不去重。"""
    inv1 = _make_invoice("A001", "10001")
    inv2 = _make_invoice("A001", "10001")
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        path = f.name
    try:
        removed = Exporter().export([inv1, inv2], path, ExportMode.SUMMARY, dedup=False)
        ws = openpyxl.load_workbook(path).active
        assert ws.max_row == 3  # header + 2 rows
        assert removed == 0
    finally:
        os.unlink(path)


def test_deduplicate_no_duplicates():
    """无重复时 removed=0。"""
    invoices = [_make_invoice("A001", "10001"), _make_invoice("B002", "20002")]
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        path = f.name
    try:
        removed = Exporter().export(invoices, path, ExportMode.SUMMARY)
        assert removed == 0
        assert openpyxl.load_workbook(path).active.max_row == 3
    finally:
        os.unlink(path)
